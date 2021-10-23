# Author: Karthik Reddy Pagilla

import datetime
from openpyxl import Workbook

class City:
    def __init__(self, name, h):
        self.name = name
        self.neighbours = None
        self.h = h
        self.visited = None

    def toString(self):
        print("City: " + self.name)
        print("Straight Line Distance to Bucharest: " + str(self.h))
        print()
        
        for neighbour in self.neighbours:
            print("Neigbour: ")
            print("Name: " + neighbour.name)
            print("Distance: " + str(self.neighbours[neighbour]))
            print()
            print("========================================")
            print()
        print("*********************************************************\n")

class Node:
    def __init__(self, city):
        self.city = city
        self.children = []
        self.parent = None
        self.depth = 0
        self.path_cost = 0

    def toString():
        print("The Path Cost: " + str(self.path_cost))

nodes_visited = 0

allCitiesList = []
allCitiesHash = {}

def reset():
    global allCitiesHash
    global allCitiesList
    
    Arad = City("Arad", 366)
    Bucharest = City("Bucharest", 0)
    Craiova = City("Craiova", 160)
    Dobreta = City("Dobreta", 242)
    Eforie = City("Eforie", 161)
    Fagaras = City("Fagaras", 176)
    Giurgiu = City("Giurgiu", 77)
    Hirsova = City("Hirsova", 151)
    Iasi = City("Iasi", 226)
    Lugoj = City("Lugoj", 244)
    Mehadia = City("Mehadia", 241)
    Neamt = City("Neamt", 234)
    Oradea = City("Oradea", 380)
    Pitesti = City("Pitesti", 100)
    Rimnicu_Vilcea = City("Rimnicu_Vilcea", 193)
    Sibiu = City("Sibiu", 253)
    Timisoara = City("Timisoara", 329)
    Urziceni = City("Urziceni", 80)
    Vaslui = City("Vaslui", 199)
    Zerind = City("Zerind", 374)

    Arad.neighbours = {Zerind:75, Timisoara:118, Sibiu:140}
    Zerind.neighbours = {Arad:75, Oradea:71}
    Oradea.neighbours = {Zerind:71, Sibiu:151}
    Timisoara.neighbours = {Arad:118, Lugoj:111}
    Lugoj.neighbours = {Timisoara:111, Mehadia:70}
    Mehadia.neighbours = {Lugoj:70, Dobreta:75}
    Dobreta.neighbours = {Mehadia:75, Craiova:120}
    Craiova.neighbours = {Dobreta:120, Rimnicu_Vilcea:146, Pitesti:138}
    Rimnicu_Vilcea.neighbours = {Craiova:146, Pitesti:97, Sibiu:80}
    Sibiu.neighbours = {Rimnicu_Vilcea:80, Fagaras:99, Oradea:151, Arad:140}
    Fagaras.neighbours = {Sibiu:99, Bucharest:211}
    Pitesti.neighbours = {Rimnicu_Vilcea:97, Craiova:138, Bucharest:101}
    Bucharest.neighbours = {Giurgiu:90, Urziceni:85, Pitesti:101, Fagaras:211}
    Urziceni.neighbours = {Bucharest:85, Hirsova:98, Vaslui:142}
    Hirsova.neighbours = {Urziceni:98, Eforie:86}
    Eforie.neighbours = {Hirsova:86}
    Vaslui.neighbours = {Urziceni:142, Iasi:92}
    Iasi.neighbours = {Vaslui:92, Neamt:87}
    Neamt.neighbours = {Iasi:87}
    Giurgiu.neighbours = {Bucharest:90}

    allCitiesList = [Arad, Bucharest, Craiova, Dobreta, Eforie, Fagaras, Giurgiu, Hirsova, Iasi, Lugoj, Mehadia, Neamt, Oradea, Pitesti, Rimnicu_Vilcea, Sibiu, Timisoara, Urziceni, Vaslui, Zerind]
    allCitiesHash = {"Arad":Arad, "Bucharest":Bucharest, "Craiova":Craiova, "Dobreta":Dobreta,
                    "Eforie":Eforie, "Fagaras":Fagaras, "Giurgiu":Giurgiu, "Hirsova":Hirsova, "Iasi":Iasi, 
                    "Lugoj":Lugoj, "Mehadia":Mehadia, "Neamt":Neamt, "Oradea":Oradea, "Pitesti":Pitesti, 
                    "Rimnicu_Vilcea":Rimnicu_Vilcea, "Sibiu":Sibiu, "Timisoara":Timisoara, "Urziceni":Urziceni, "Vaslui":Vaslui, "Zerind":Zerind}

reset()

# Method to test the implementation of the Graph
def mapTest():
    global allCitiesList
    
    edge_list = []

    for city in allCitiesList:
        for neighbour in city.neighbours.keys():
            edge_list.append((city.name, neighbour.name, city.neighbours[neighbour]))

    for edge in edge_list:
        print(edge)

def allCitiesFromList():
    global allCitiesList
    
    city_names = []
    for city in allCitiesList:
        city_names.append(city.name)
    return city_names

def allCitiesFromHTable():
    global allCitiesHash
    
    cities = []
    for city in allCitiesHash.keys():
        cities.append(allCitiesHash[city])
    return cities

def getCityFromList(name):
    global allCitiesList

    for city in allCitiesList:
        if city.name == name:
            return city

def getCityFromHtable(name):
    global allCitiesHash
    
    return allCitiesHash[name]

def neighboursUsingList(name):
    city = getCityFromList(name)

    return city.neighbours.keys()

def neighboursUsingHtable(name):
    city = getCityFromHtable(name)

    return city.neighbours.keys()

def neighborsWithinD(myCity, distance):
    city = getCityFromHtable(myCity)

    valid_neighbours = {}

    for neighbour in city.neighbours.keys():
        if city.neighbours[neighbour] <= distance:
            valid_neighbours[neighbour] = city.neighbours[neighbour]

    return valid_neighbours

def neighborsP(cityOne, cityTwo):
    city_one = getCityFromHtable(cityOne)
    city_two = getCityFromHtable(cityTwo)

    for neighbour in city_one.neighbours.keys():
        if neighbour.name == city_two.name:
            return city_one.neighbours[neighbour]

    return None

def expand_node(node, typeOfSearch):
    global nodes_visited
    
    neighbors = node.city.neighbours.keys()

    if (typeOfSearch == "Graph"):
        for neighbor in neighbors:
            if neighbor.visited == None:
                child = Node(neighbor)
                child.parent = node
                child.depth = node.depth + 1
                child.path_cost = node.path_cost + node.city.neighbours[neighbor]
                node.children.append(child)
        node.city.visited = "T"
    else:
        for neighbor in neighbors:
            child = Node(neighbor)
            child.parent = node
            child.depth = node.depth + 1
            child.path_cost = node.path_cost + node.city.neighbours[neighbor]
            node.children.append(child) 

    nodes_visited += 1
    return node

def evaluate_node(node, strategy):
    if (strategy == "Greedy"):
        return node.city.h
    elif (strategy == "UCS"):
        return node.path_cost
    elif (strategy == "A*"):
        return node.city.h + node.path_cost

def node_to_expand(fringe, strategy):
    min_value = 10000000
    expand_node = None
    for node in fringe:
        if evaluate_node(node, strategy) < min_value:
            min_value = evaluate_node(node, strategy)
            expand_node = node
    
    return expand_node

def general_search(name, strategy, typeOfSearch):
    global nodes_visited
    nodes_visited = 0

    city = getCityFromHtable(name)

    node = Node(city)
    node.path_cost = 0
    node.depth = 0
    node.parent = None

    fringe = []
    
    while(node != None):
        if (node.city.name == "Bucharest"):
            return node

        node = expand_node(node, typeOfSearch)

        parents = []

        temp_node = node

        if typeOfSearch == "Tree":
            while (temp_node.parent != None):
                parents.append(temp_node.parent.city)
                temp_node = temp_node.parent

        for child in node.children:
            if child.city not in parents:
                fringe.append(child)

        node = node_to_expand(fringe, strategy)
        fringe.remove(node)

    return None

def path_finder(node):
    path = []
    while(node != None):
        path.append(node.city.name)
        node = node.parent

    path.reverse()
    result = ""
    for city in path:
        if city == "Bucharest":
            result = result + city
        else:
            result = result + city + " --> "
    
    return result

def result_generator():
    
    result = None

    workbook = Workbook()
    sheet = workbook.active

    # UCS - Graph Search

    # print("*"*50)
    # print("*",end="")
    # print(" "*10, end="")
    # print("UCS - Graph Search", end="")
    # print(" "*20, end="")
    # print("*")
    # print("*"*50)
    # print()

    # n = 1
    # for city in allCitiesFromList():

    #     a = datetime.datetime.now()
    #     result = general_search(city, "UCS", "Graph")
    #     b = datetime.datetime.now()

    #     diff = b - a
    #     cpu_time = diff.total_seconds() * 1000

    #     if result != None:
    #         c1 = sheet.cell(row = n, column = 1)
    #         c1.value = path_finder(result)
    #         print(path_finder(result))

    #         c2 = sheet.cell(row = n, column = 2)
    #         c2.value = nodes_visited
    #         print("Nodes Visited: " + str(nodes_visited))

    #         c3 = sheet.cell(row = n, column = 3)
    #         c3.value = result.path_cost
    #         print("Cost of Path: " + str(result.path_cost))

    #         c4 = sheet.cell(row = n, column = 4)
    #         c4.value = cpu_time
    #         print("CPU Time (in ms): " + str(cpu_time))

    #         print()
    #         print("="*20)
    #         print()

    #         n = n + 1

    #     reset()

    # UCS - Tree Search

    # print("*"*50)
    # print("*",end="")
    # print(" "*10, end="")
    # print("UCS - Tree Search", end="")
    # print(" "*21, end="")
    # print("*")
    # print("*"*50)
    # print()

    # n = 1
    # for city in allCitiesFromList():

    #     a = datetime.datetime.now()
    #     result = general_search(city, "UCS", "Tree")
    #     b = datetime.datetime.now()

    #     diff = b - a
    #     cpu_time = diff.total_seconds() * 1000

    #     if result != None:
    #         c1 = sheet.cell(row = n, column = 1)
    #         c1.value = path_finder(result)
    #         print(path_finder(result))

    #         c2 = sheet.cell(row = n, column = 2)
    #         c2.value = nodes_visited
    #         print("Nodes Visited: " + str(nodes_visited))

    #         c3 = sheet.cell(row = n, column = 3)
    #         c3.value = result.path_cost
    #         print("Cost of Path: " + str(result.path_cost))

    #         c4 = sheet.cell(row = n, column = 4)
    #         c4.value = cpu_time
    #         print("CPU Time (in ms): " + str(cpu_time))

    #         print()
    #         print("="*20)
    #         print()

    #         n = n + 1

    #     reset()

    # Greedy - Graph Search

    # print("*"*50)
    # print("*",end="")
    # print(" "*10, end="")
    # print("Greedy - Graph Search", end="")
    # print(" "*17, end="")
    # print("*")
    # print("*"*50)
    # print()

    # n = 1
    # for city in allCitiesFromList():

    #     a = datetime.datetime.now()
    #     result = general_search(city, "Greedy", "Graph")
    #     b = datetime.datetime.now()

    #     diff = b - a
    #     cpu_time = diff.total_seconds() * 1000

    #     if result != None:
    #         c1 = sheet.cell(row = n, column = 1)
    #         c1.value = path_finder(result)
    #         print(path_finder(result))

    #         c2 = sheet.cell(row = n, column = 2)
    #         c2.value = nodes_visited
    #         print("Nodes Visited: " + str(nodes_visited))

    #         c3 = sheet.cell(row = n, column = 3)
    #         c3.value = result.path_cost
    #         print("Cost of Path: " + str(result.path_cost))

    #         c4 = sheet.cell(row = n, column = 4)
    #         c4.value = cpu_time
    #         print("CPU Time (in ms): " + str(cpu_time))

    #         print()
    #         print("="*20)
    #         print()

    #         n = n + 1

    #     reset()

    # Greedy - Tree Search

    # print("*"*50)
    # print("*",end="")
    # print(" "*10, end="")
    # print("Greedy - Tree Search", end="")
    # print(" "*18, end="")
    # print("*")
    # print("*"*50)
    # print()

    # n = 1
    # for city in allCitiesFromList():

    #     a = datetime.datetime.now()
    #     result = general_search(city, "Greedy", "Tree")
    #     b = datetime.datetime.now()

    #     diff = b - a
    #     cpu_time = diff.total_seconds() * 1000

    #     if result != None:
    #         c1 = sheet.cell(row = n, column = 1)
    #         c1.value = path_finder(result)
    #         print(path_finder(result))

    #         c2 = sheet.cell(row = n, column = 2)
    #         c2.value = nodes_visited
    #         print("Nodes Visited: " + str(nodes_visited))

    #         c3 = sheet.cell(row = n, column = 3)
    #         c3.value = result.path_cost
    #         print("Cost of Path: " + str(result.path_cost))

    #         c4 = sheet.cell(row = n, column = 4)
    #         c4.value = cpu_time
    #         print("CPU Time (in ms): " + str(cpu_time))

    #         print()
    #         print("="*20)
    #         print()

    #         n = n + 1

    #     reset()

    # A* - Graph Search

    # print("*"*50)
    # print("*",end="")
    # print(" "*10, end="")
    # print("A* - Graph Search", end="")
    # print(" "*21, end="")
    # print("*")
    # print("*"*50)
    # print()

    # n = 1
    # for city in allCitiesFromList():

    #     a = datetime.datetime.now()
    #     result = general_search(city, "A*", "Graph")
    #     b = datetime.datetime.now()

    #     diff = b - a
    #     cpu_time = diff.total_seconds() * 1000

    #     if result != None:
    #         c1 = sheet.cell(row = n, column = 1)
    #         c1.value = path_finder(result)
    #         print(path_finder(result))

    #         c2 = sheet.cell(row = n, column = 2)
    #         c2.value = nodes_visited
    #         print("Nodes Visited: " + str(nodes_visited))

    #         c3 = sheet.cell(row = n, column = 3)
    #         c3.value = result.path_cost
    #         print("Cost of Path: " + str(result.path_cost))

    #         c4 = sheet.cell(row = n, column = 4)
    #         c4.value = cpu_time
    #         print("CPU Time (in ms): " + str(cpu_time))

    #         print()
    #         print("="*20)
    #         print()

    #         n = n + 1

    #     reset()

    # A* - Tree Search

    print("*"*50)
    print("*",end="")
    print(" "*10, end="")
    print("A* - Tree Search", end="")
    print(" "*22, end="")
    print("*")
    print("*"*50)
    print()

    n = 1
    for city in allCitiesFromList():

        a = datetime.datetime.now()
        result = general_search(city, "A*", "Tree")
        b = datetime.datetime.now()

        diff = b - a
        cpu_time = diff.total_seconds() * 1000

        if result != None:
            c1 = sheet.cell(row = n, column = 1)
            c1.value = path_finder(result)
            print(path_finder(result))

            c2 = sheet.cell(row = n, column = 2)
            c2.value = nodes_visited
            print("Nodes Visited: " + str(nodes_visited))

            c3 = sheet.cell(row = n, column = 3)
            c3.value = result.path_cost
            print("Cost of Path: " + str(result.path_cost))

            c4 = sheet.cell(row = n, column = 4)
            c4.value = cpu_time
            print("CPU Time (in ms): " + str(cpu_time))

            print()
            print("="*20)
            print()

            n = n + 1

        reset()

    #workbook.save("/content/holiday5.xlsx")

result_generator()